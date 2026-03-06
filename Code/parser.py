import pandas as pd
import openpyxl
from pathlib import Path


def list_sheets(filepath: str):
    """List all sheet names in the Excel file."""
    workbook = openpyxl.load_workbook(filepath)
    print("Available sheets:")
    for i, sheet_name in enumerate(workbook.sheetnames, 1):
        print(f"  {i}. {sheet_name}")
    return workbook.sheetnames


def parse_iets_model(filepath: str) -> dict:
    """
    Parse IETS_ModelName_v6.xlsx and extract all sheets.
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        Dictionary with all sheet DataFrames
    """
    excel_file = Path(filepath)
    
    if not excel_file.exists():
        raise FileNotFoundError(f"File not found: {filepath}")
    
    # Get all sheet names
    workbook = openpyxl.load_workbook(excel_file)
    sheet_names = workbook.sheetnames
    
    result = {}
    
    # Read all sheets
    for sheet_name in sheet_names:
        try:
            result[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
            print(f"✓ Loaded '{sheet_name}' sheet: {result[sheet_name].shape}")
        except Exception as e:
            print(f"⚠ Could not load '{sheet_name}' sheet: {e}")
    
    return result


def main():
    # Path to the Excel file
    filepath = Path(__file__).parent.parent / "Base" / "IETS_ModelName_v6.xlsx"
    
    # Parse the file
    data = parse_iets_model(str(filepath))
    
    # Display summary
    for sheet_name, df in data.items():
        print(f"\n=== {sheet_name} ===")
        print(f"Shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        print(df.head())
    
    return data


if __name__ == "__main__":
    data = main()