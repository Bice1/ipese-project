"""
IETS v6 Excel Template Parser
Parses IETS_ModelName_v6.xlsx files and prints a structured summary to console.
"""

from __future__ import annotations

import sys
import os
from dataclasses import dataclass, field
from typing import Any

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FIXED_SHEETS = frozenset([
    "README", "METADATA", "CONNECTORS", "VARIABLES",
    "CALCULATIONS", "CHANGELOG", "SUPPL MATERIAL",
])

SECTION_LABELS = {
    "connectors":   "CONNECTORS",
    "heat_streams": "HEAT STREAMS",
    "equipments":   "EQUIPMENTS",
    "opex":         "OPERATING COST",
}

CONNECTOR_COL_MAP = {
    "NAME (ALIAS)":            "name",
    "TYPE":                    "type",
    "DIRECTION":               "direction",
    "FLOW VALUE":              "flow_value",
    "FLOW UNIT":               "flow_unit",
    "TEMPERATURE VALUE":       "T_value",
    "TEMPERATURE UNIT":        "T_unit",
    "PRESSURE VALUE":          "P_value",
    "PRESSURE UNIT":           "P_unit",
    "VAPOR FRACTION [MASS]":   "vapor_fraction",
    "PHASE":                   "phase",
    "COMPOSITION [MOLE/MASS]": "composition",
    "REFERENCE STREAM":        "reference_stream",
    "DESCRIPTION":             "description",
}

HEAT_STREAM_COL_MAP = {
    "NAME":                       "name",
    "TYPE":                       "type",
    "INLET TEMPERATURE VALUE":    "T_in",
    "INLET TEMPERATURE UNIT":     "T_in_unit",
    "OUTLET TEMPERATURE VALUE":   "T_out",
    "OUTLET TEMPERATURE UNIT":    "T_out_unit",
    "INLET ENTHALPY FLOW VALUE":  "H_in",
    "INLET ENTHALPY FLOW UNIT":   "H_in_unit",
    "OUTLET ENTHALPY FLOW VALUE": "H_out",
    "OUTLET ENTHALPY FLOW UNIT":  "H_out_unit",
    "SPEC HEAT CAP VALUE":        "cp",
    "SPEC HEAT CAP UNIT":         "cp_unit",
    "MASS FLOW VALUE":            "mdot",
    "MASS FLOW UNIT":             "mdot_unit",
    "DTMIN CONTR VALUE":          "dtmin_contr",
    "DTMIN CONTR UNIT":           "dtmin_unit",
    "HEAT TRANSFER COEF VALUE":   "alpha",
    "HEAT TRANSFER COEF UNIT":    "alpha_unit",
    "HEAT CASCADE":               "heat_cascade",
}


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class UnitHeader:
    name: str
    fmin: float | None
    fmax: float | None


@dataclass
class UnitData:
    name: str
    fmin: float | None
    fmax: float | None
    connectors: pd.DataFrame
    heat_streams: pd.DataFrame
    equipments: list[dict]
    description: str = ""


@dataclass
class ParsedModel:
    filepath: str
    metadata: dict[str, Any]
    unit_names: list[str]
    units: dict[str, UnitData]
    external_connectors: pd.DataFrame
    internal_connectors: pd.DataFrame
    global_uid_registry: list[str]
    variables: pd.DataFrame
    calculations: pd.DataFrame | None
    warnings: list[str] = field(default_factory=list)
    category: str = ""


# ---------------------------------------------------------------------------
# Parser class
# ---------------------------------------------------------------------------

class IETSParser:

    def __init__(self, filepath: str) -> None:
        self.filepath = filepath
        self.warnings: list[str] = []

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def parse(self) -> ParsedModel:
        wb = self._load_workbook()
        unit_names = self._get_unit_sheet_names(wb)

        metadata = self._parse_metadata(wb["METADATA"])

        conn_data = self._parse_connectors_sheet(wb["CONNECTORS"])

        variables = pd.DataFrame()
        if "VARIABLES" in wb.sheetnames:
            try:
                variables = self._parse_variables(wb["VARIABLES"])
            except Exception as e:
                self._log_warning(f"VARIABLES sheet parse failed: {e}")

        calculations = None
        if "CALCULATIONS" in wb.sheetnames:
            try:
                calculations = self._parse_calculations(wb["CALCULATIONS"])
            except Exception as e:
                self._log_warning(f"CALCULATIONS sheet parse failed: {e}")

        units: dict[str, UnitData] = {}
        for name in unit_names:
            try:
                units[name] = self._parse_unit_sheet(wb[name])
            except Exception as e:
                self._log_warning(f"Unit sheet '{name}' parse failed: {e}")

        return ParsedModel(
            filepath=self.filepath,
            metadata=metadata,
            unit_names=unit_names,
            units=units,
            external_connectors=conn_data.get("external", pd.DataFrame()),
            internal_connectors=conn_data.get("internal", pd.DataFrame()),
            global_uid_registry=conn_data.get("global_registry", []),
            variables=variables,
            calculations=calculations,
            warnings=self.warnings,
        )

    # ------------------------------------------------------------------
    # Workbook helpers
    # ------------------------------------------------------------------

    def _load_workbook(self) -> openpyxl.Workbook:
        return openpyxl.load_workbook(self.filepath, data_only=True)

    def _get_unit_sheet_names(self, wb: openpyxl.Workbook) -> list[str]:
        return [s for s in wb.sheetnames if s not in FIXED_SHEETS]

    # ------------------------------------------------------------------
    # METADATA sheet
    # ------------------------------------------------------------------

    def _parse_metadata(self, ws) -> dict[str, Any]:
        metadata: dict[str, Any] = {}
        consecutive_empty = 0
        for row in ws.iter_rows(values_only=True):
            key = row[0]
            val = row[1] if len(row) > 1 else None
            if key is None:
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    break
                continue
            consecutive_empty = 0
            key_str = str(key).strip()
            if key_str:
                metadata[key_str] = val
        return metadata

    # ------------------------------------------------------------------
    # CONNECTORS sheet
    # ------------------------------------------------------------------

    def _parse_connectors_sheet(self, ws) -> dict:
        rows = list(ws.iter_rows(values_only=True))

        # Find EXTERNAL CONNECTORS and INTERNAL CONNECTORS section labels
        ext_start = None
        int_start = None
        for i, row in enumerate(rows):
            col_a = str(row[0]).strip().upper() if row[0] is not None else ""
            if col_a == "EXTERNAL CONNECTORS":
                ext_start = i
            elif col_a == "INTERNAL CONNECTORS":
                int_start = i

        external = pd.DataFrame()
        internal = pd.DataFrame()

        if ext_start is not None:
            h_idx = self._next_non_empty(rows, ext_start + 1)
            stop = int_start if int_start is not None else len(rows)
            if h_idx is not None:
                external = self._parse_table_generic(rows, h_idx, stop, CONNECTOR_COL_MAP)

        if int_start is not None:
            h_idx = self._next_non_empty(rows, int_start + 1)
            if h_idx is not None:
                internal = self._parse_table_generic(rows, h_idx, len(rows), CONNECTOR_COL_MAP)

        # Global UID registry — col Q (index 16)
        global_registry: list[str] = []
        registry_started = False
        for row in rows:
            if len(row) <= 16:
                continue
            val = row[16]
            if val is None:
                continue
            val_str = str(val).strip()
            if not registry_started:
                if val_str.upper() in ("UNIQUE IDENTIFIER NUMBER, UID", "UID"):
                    registry_started = True
                    continue
                registry_started = True
            if val_str:
                global_registry.append(val_str)

        return {
            "external": external,
            "internal": internal,
            "global_registry": global_registry,
        }

    # ------------------------------------------------------------------
    # VARIABLES sheet
    # ------------------------------------------------------------------

    def _parse_variables(self, ws) -> pd.DataFrame:
        rows = list(ws.iter_rows(values_only=True))
        # Find header row containing "NAME" and "TYPE"
        h_idx = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            if "NAME" in cells and "TYPE" in cells:
                h_idx = i
                break
        if h_idx is None:
            return pd.DataFrame()
        return self._parse_table_generic(rows, h_idx, len(rows), {})

    # ------------------------------------------------------------------
    # CALCULATIONS sheet
    # ------------------------------------------------------------------

    def _parse_calculations(self, ws) -> pd.DataFrame:
        rows = list(ws.iter_rows(values_only=True))
        h_idx = self._next_non_empty(rows, 0)
        if h_idx is None:
            return pd.DataFrame()
        return self._parse_table_generic(rows, h_idx, len(rows), {})

    # ------------------------------------------------------------------
    # Unit sheet parsing
    # ------------------------------------------------------------------

    def _parse_unit_sheet(self, ws) -> UnitData:
        rows = list(ws.iter_rows(values_only=True))

        header = self._find_unit_header(rows)

        # Locate sections
        _, conn_h = self._find_section(rows, SECTION_LABELS["connectors"])
        hs_sec, hs_h = self._find_section(rows, SECTION_LABELS["heat_streams"])
        eq_sec, eq_h = self._find_section(rows, SECTION_LABELS["equipments"])
        opex_sec, _ = self._find_section(rows, SECTION_LABELS["opex"])

        # Determine stop rows for each section
        def first_of(*candidates):
            valid = [c for c in candidates if c is not None]
            return min(valid) if valid else len(rows)

        conn_stop = first_of(hs_sec, eq_sec, opex_sec)
        hs_stop = first_of(eq_sec, opex_sec)
        eq_stop = opex_sec if opex_sec is not None else len(rows)

        # Parse each table
        connectors = pd.DataFrame()
        if conn_h is not None:
            connectors = self._parse_table_generic(rows, conn_h, conn_stop, CONNECTOR_COL_MAP)

        heat_streams = pd.DataFrame()
        if hs_h is not None:
            heat_streams = self._parse_heat_streams_table(rows, hs_h, hs_stop)

        equipments: list[dict] = []
        if eq_h is not None:
            equipments = self._parse_equipments_table(rows, eq_h, eq_stop)

        return UnitData(
            name=header.name,
            fmin=header.fmin,
            fmax=header.fmax,
            connectors=connectors,
            heat_streams=heat_streams,
            equipments=equipments,
        )

    def _find_unit_header(self, rows: list[tuple]) -> UnitHeader:
        name = None
        fmin = None
        fmax = None

        for i, row in enumerate(rows):
            col_a = str(row[0]).strip().upper() if row[0] is not None else ""
            if col_a == "UNIT NAME":
                j = self._next_non_empty(rows, i + 1)
                if j is not None:
                    name = str(rows[j][0]).strip() if rows[j][0] is not None else None
                    fmin = self._safe_numeric(rows[j][1]) if len(rows[j]) > 1 else None
                    fmax = self._safe_numeric(rows[j][2]) if len(rows[j]) > 2 else None
                break

        if name is None:
            name = "Unknown"
        return UnitHeader(name=name, fmin=fmin, fmax=fmax)

    # ------------------------------------------------------------------
    # Section detection utilities
    # ------------------------------------------------------------------

    def _find_section(self, rows: list[tuple], label: str) -> tuple[int | None, int | None]:
        label_upper = label.strip().upper()
        for i, row in enumerate(rows):
            col_a = str(row[0]).strip().upper() if row[0] is not None else ""
            col_b = row[1] if len(row) > 1 else None
            if col_a == label_upper and col_b is None:
                h_idx = self._next_non_empty(rows, i + 1)
                return (i, h_idx)
        return (None, None)

    def _next_non_empty(self, rows: list[tuple], start: int) -> int | None:
        for i in range(start, len(rows)):
            if any(c is not None for c in rows[i]):
                return i
        return None

    # ------------------------------------------------------------------
    # Generic table parser
    # ------------------------------------------------------------------

    def _parse_table_generic(
        self,
        rows: list[tuple],
        header_idx: int,
        stop_idx: int,
        col_renames: dict[str, str],
    ) -> pd.DataFrame:
        header_row = rows[header_idx]

        col_names = []
        for cell in header_row:
            if cell is None:
                col_names.append(None)
                continue
            cell_upper = str(cell).strip().upper()
            matched = None
            for k, v in col_renames.items():
                if k.strip().upper() == cell_upper:
                    matched = v
                    break
            col_names.append(matched if matched else str(cell).strip())

        data = []
        for row in rows[header_idx + 1: stop_idx]:
            if all(c is None for c in row):
                continue
            row_dict = {}
            for j, col in enumerate(col_names):
                if col is None:
                    continue
                val = row[j] if j < len(row) else None
                row_dict[col] = val
            if any(v is not None for v in row_dict.values()):
                data.append(row_dict)

        return pd.DataFrame(data)

    # ------------------------------------------------------------------
    # Heat streams parser
    # ------------------------------------------------------------------

    def _parse_heat_streams_table(
        self, rows: list[tuple], header_idx: int, stop_idx: int
    ) -> pd.DataFrame:
        df = self._parse_table_generic(rows, header_idx, stop_idx, HEAT_STREAM_COL_MAP)
        if df.empty:
            return df

        if "type" in df.columns:
            df["type_raw"] = df["type"]
            df["type"] = df["type"].apply(
                lambda v: str(v).strip().title() if v is not None else v
            )

        def is_phase_change(row):
            t_in = row.get("T_in")
            t_out = row.get("T_out")
            try:
                return abs(float(t_in) - float(t_out)) < 1.0
            except (TypeError, ValueError):
                return False

        df["is_phase_change"] = df.apply(is_phase_change, axis=1)
        return df

    # ------------------------------------------------------------------
    # Equipments parser (multi-row grouping)
    # ------------------------------------------------------------------

    def _parse_equipments_table(
        self, rows: list[tuple], header_idx: int, stop_idx: int
    ) -> list[dict]:
        header_row = rows[header_idx]
        col_names = [
            str(c).strip().lower() if c is not None else None
            for c in header_row
        ]

        def get_col(row, name):
            try:
                idx = next(i for i, c in enumerate(col_names) if c == name)
                return row[idx] if idx < len(row) else None
            except StopIteration:
                return None

        equipments: list[dict] = []
        current: dict | None = None

        for row in rows[header_idx + 1: stop_idx]:
            if all(c is None for c in row):
                continue

            col_a = row[0] if len(row) > 0 else None

            if col_a is not None:
                if current is not None:
                    equipments.append(current)
                current = {
                    "name":      str(col_a).strip(),
                    "type":      get_col(row, "type"),
                    "subtype":   get_col(row, "subtype"),
                    "reference": get_col(row, "reference"),
                    "params":    [],
                }
                param = get_col(row, "params")
                value = get_col(row, "value")
                unit  = get_col(row, "unit")
                if any(v is not None for v in [param, value, unit]):
                    current["params"].append({"param": param, "value": value, "unit": unit})
            else:
                if current is None:
                    continue
                param = get_col(row, "params")
                value = get_col(row, "value")
                unit  = get_col(row, "unit")
                if any(v is not None for v in [param, value, unit]):
                    current["params"].append({"param": param, "value": value, "unit": unit})

        if current is not None:
            equipments.append(current)

        return equipments

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _safe_numeric(self, val: Any) -> float | str | None:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        if s.upper() in ("SAT", "-", "N/A", ""):
            return s if s else None
        try:
            return float(s)
        except (ValueError, TypeError):
            return s

    def _log_warning(self, msg: str) -> None:
        self.warnings.append(msg)

    # ------------------------------------------------------------------
    # Print summary
    # ------------------------------------------------------------------

    def print_summary(self, model: ParsedModel) -> None:
        m = model.metadata
        sep = "=" * 80

        print(sep)
        print("IETS MODEL PARSER SUMMARY")
        print(sep)
        print(f"File:        {os.path.basename(model.filepath)}")
        print(f"Model UID:   {m.get('MODEL UID', 'N/A')}")
        print(f"Model Name:  {m.get('MODEL NAME', 'N/A')}")
        print(f"Authors:     {m.get('AUTHORS AND CONTRIBUTORS', 'N/A')}")
        print(f"Grade:       {m.get('GRADE', 'N/A')}")
        print(f"TRL:         {m.get('TRL', 'N/A')}")
        print(f"Sharing:     Layer {m.get('SHARING LAYER', '?')} | {m.get('CONFIDENTIALITY', 'N/A')}")
        print(f"Software:    {m.get('SOFTWARE', 'N/A')}")
        print(f"Keywords:    {m.get('KEYWORDS', 'N/A')}")
        print(f"Created:     {m.get('CREATION DATE', 'N/A')}  |  Updated: {m.get('LAST UPDATED', 'N/A')}")
        print(f"Description: {m.get('DESCRIPTION', 'N/A')}")
        print()
        print(f"Unit Sheets Found: {', '.join(model.unit_names) or '(none)'}")
        print(
            f"External Connectors: {len(model.external_connectors)}  |  "
            f"Internal Connectors: {len(model.internal_connectors)}  |  "
            f"Global UID Registry: {len(model.global_uid_registry)} entries"
        )

        for unit_name in model.unit_names:
            if unit_name not in model.units:
                continue
            unit = model.units[unit_name]
            print()
            print(sep)
            scalable = (unit.fmin != unit.fmax) if (unit.fmin is not None and unit.fmax is not None) else None
            scale_str = "SCALABLE" if scalable else "FIXED"
            print(f"UNIT: {unit.name}")
            print(f"  Capacity multipliers: Fmin={unit.fmin}, Fmax={unit.fmax}  [{scale_str}]")

            # Connectors
            print(f"\n  Connectors ({len(unit.connectors)}):")
            if not unit.connectors.empty:
                for _, r in unit.connectors.iterrows():
                    name     = r.get("name", "")
                    direction = r.get("direction", "")
                    ctype    = r.get("type", "")
                    flow_val = r.get("flow_value", "")
                    flow_unit = r.get("flow_unit", "")
                    t_val    = r.get("T_value", "")
                    t_unit   = r.get("T_unit", "")
                    flow_str = f"{flow_val} {flow_unit}".strip() if flow_val is not None else "-"
                    t_str    = f"{t_val} {t_unit}".strip() if t_val is not None else "-"
                    print(f"    {str(name):<22} {str(direction):<5} {str(ctype):<12} flow={flow_str:<16} T={t_str}")

            # Heat streams
            print(f"\n  Heat Streams ({len(unit.heat_streams)}):")
            if not unit.heat_streams.empty:
                cascades = set()
                for _, r in unit.heat_streams.iterrows():
                    name     = r.get("name", "")
                    stype    = r.get("type", "")
                    t_in     = r.get("T_in", "")
                    t_out    = r.get("T_out", "")
                    h_in     = r.get("H_in", "")
                    h_out    = r.get("H_out", "")
                    cascade  = r.get("heat_cascade", "")
                    phase_chg = r.get("is_phase_change", False)
                    if cascade:
                        cascades.add(str(cascade))
                    print(
                        f"    {str(name):<30} {str(stype):<5} "
                        f"T: {str(t_in):<7}-> {str(t_out):<9}"
                        f"H: {str(h_in):<7}-> {str(h_out):<7}"
                        f"cascade={str(cascade):<15} phase_chg={phase_chg}"
                    )
                print(f"\n  Heat Cascade Labels: {', '.join(sorted(cascades)) or '(none)'}")

            # Equipments
            print(f"\n  Equipments ({len(unit.equipments)}):")
            for eq in unit.equipments:
                eq_type  = eq.get("type", "-")
                eq_sub   = eq.get("subtype", "-")
                n_params = len(eq.get("params", []))
                print(f"    {str(eq.get('name', '')):<28} [{eq_type} / {eq_sub}] — {n_params} params")

        # Warnings
        print()
        print(sep)
        print(f"WARNINGS ({len(model.warnings)}):")
        if model.warnings:
            for w in model.warnings:
                print(f"  ! {w}")
        else:
            print("  (none)")
        print(sep)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(script_dir))
    default_file = os.path.join(project_root, "Base", "IETS_ModelName_v6.xlsx")

    filepath = sys.argv[1] if len(sys.argv) > 1 else default_file

    if not os.path.exists(filepath):
        print(f"Error: file not found: {filepath}")
        sys.exit(1)

    parser = IETSParser(filepath)
    model = parser.parse()
    parser.print_summary(model)
