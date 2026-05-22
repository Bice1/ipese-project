import json
import tempfile
import pandas as pd
import re
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from fastapi import HTTPException, status

# =====================
# USER CONFIG: Set filenames here
# =====================
INPUT_EXCEL_NAME = "CCSAdsorption_v6.xlsx"
OUTPUT_JSON_NAME = "CCSAdsorption_v6.json"
V6_EXCEL_DIR = "v6_excel"   # folder containing all v6 Excel files
V6_JSON_DIR = "v6_json"     # output folder for generated JSON files
# =====================
# END USER CONFIG
# =====================


class FormulaEvaluator:
    """Recursive formula evaluator for Excel formulas."""

    def __init__(self, formula_ws):
        self.formula_ws = formula_ws
        self.cache = {}  # Cache calculated values
        self.evaluating = set()  # Track cells being evaluated (prevent infinite recursion)

    def get_cell_value(self, cell_ref):
        """Get the value of a cell, evaluating formulas if necessary."""
        if cell_ref in self.cache:
            return self.cache[cell_ref]

        if cell_ref in self.evaluating:
            # Circular reference detected
            return None

        # Parse cell reference
        col_letter = ''.join([c for c in cell_ref if c.isalpha()])
        row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
        col_num = column_index_from_string(col_letter)

        # Get the cell value
        cell_value = self.formula_ws.cell(row=row_num, column=col_num).value

        if cell_value is None:
            self.cache[cell_ref] = None
            return None

        # If it's a formula, evaluate it recursively
        if isinstance(cell_value, str) and cell_value.startswith('='):
            self.evaluating.add(cell_ref)
            try:
                result = self.evaluate_formula(cell_value)
                self.cache[cell_ref] = result
                return result
            finally:
                self.evaluating.discard(cell_ref)
        else:
            # It's a regular value
            self.cache[cell_ref] = cell_value
            return cell_value

    def evaluate_formula(self, formula):
        """Evaluate a formula by recursively resolving cell references."""
        if not formula or not formula.startswith('='):
            return None

        try:
            # Remove the = sign
            expr = formula[1:]

            # Find all cell references (like E15, E17, etc.)
            cell_refs = re.findall(r'[A-Z]+\d+', expr)

            # Replace each cell reference with its actual value
            for cell_ref in cell_refs:
                cell_value = self.get_cell_value(cell_ref)

                if cell_value is None:
                    return None

                # Replace the cell reference with the actual value
                expr = expr.replace(cell_ref, str(cell_value))

            # Evaluate the mathematical expression
            # Only allow safe operations
            allowed_chars = set('0123456789+-*/()., ')
            if all(c in allowed_chars for c in expr):
                result = eval(expr)
                return result
            else:
                return None

        except Exception:
            return None


def get_formula_and_value(formula_ws, value_ws, pandas_df, row_idx, col_idx):
    """
    Get both formula and calculated value from Excel cell by finding exact match.
    Enhanced version that calculates missing values when Excel hasn't saved them.
    """
    formula = None
    calculated_value = None

    if not formula_ws:
        return formula, calculated_value

    # Search for matching content in first column (Variable name)
    try:
        if row_idx < len(pandas_df):
            pandas_first_cell = str(pandas_df.iloc[row_idx, 0] or '').strip()
            if pandas_first_cell:
                # Search all rows in Excel for this variable name in column A
                for excel_row in range(1, formula_ws.max_row + 1):
                    excel_first_cell = str(formula_ws.cell(row=excel_row, column=1).value or '').strip()
                    if pandas_first_cell == excel_first_cell:
                        try:
                            excel_col = col_idx + 1  # Convert to 1-based indexing
                            formula_cell = formula_ws.cell(row=excel_row, column=excel_col)
                            if formula_cell.value and str(formula_cell.value).startswith('='):
                                formula = str(formula_cell.value)

                                # Try to get calculated value from Excel first
                                if value_ws:
                                    value_cell = value_ws.cell(row=excel_row, column=excel_col)
                                    calculated_value = value_cell.value

                                # If calculated value is None, try manual calculation
                                if calculated_value is None:
                                    evaluator = FormulaEvaluator(formula_ws)
                                    cell_ref = f"{get_column_letter(excel_col)}{excel_row}"
                                    manual_result = evaluator.get_cell_value(cell_ref)

                                    if manual_result is not None:
                                        calculated_value = manual_result
                                        print(f"Manual calculation: {pandas_first_cell} -> {formula} = {calculated_value}")

                                return formula, calculated_value
                        except:
                            pass
                        break
    except:
        pass

    return formula, calculated_value


def replace_cell_references_with_variable_names(formula, formula_ws, variables_ws=None):
    """Replace cell references in formulas with variable names for better readability."""
    if not formula or not formula_ws:
        return formula

    import re

    # Handle cross-sheet references like =VARIABLES!E24
    if "VARIABLES!" in formula:
        # Extract the cell reference after VARIABLES!
        variables_pattern = r'VARIABLES!([A-Z]+)(\d+)'
        variables_refs = re.findall(variables_pattern, formula)

        modified_formula = formula

        for col_letter, row_num in variables_refs:
            row_num = int(row_num)

            try:
                # Get the variable name from VARIABLES sheet column A
                if variables_ws:
                    variable_name = variables_ws.cell(row=row_num, column=1).value
                    if variable_name:
                        variable_name = str(variable_name).strip()
                        # Replace the full reference with just the variable name
                        full_ref = f"VARIABLES!{col_letter}{row_num}"
                        modified_formula = modified_formula.replace(full_ref, variable_name)
            except:
                pass

        return modified_formula

    # Handle same-sheet references (like E15, E22, etc.)
    cell_pattern = r'([A-Z]+)(\d+)'
    cell_refs = re.findall(cell_pattern, formula)

    modified_formula = formula

    for col_letter, row_num in cell_refs:
        row_num = int(row_num)

        try:
            # Get the variable name from column A of the same row
            variable_name = formula_ws.cell(row=row_num, column=1).value
            if variable_name:
                variable_name = str(variable_name).strip()
                # Replace the cell reference with the variable name
                cell_ref = f"{col_letter}{row_num}"
                modified_formula = modified_formula.replace(cell_ref, variable_name)
        except:
            pass

    return modified_formula




def excel_to_json(excel_file_path: str, output_json_path: str) -> None:
    # Load workbook for formula access
    formula_wb = load_workbook(excel_file_path, data_only=False)
    value_wb = load_workbook(excel_file_path, data_only=True)
    output: Dict[str, Any] = {
        "METADATA": {},
        "VARIABLES": [],
        "CONNECTORS": {
            "EXTERNAL CONNECTORS": [],
            "INTERNAL CONNECTORS": []
        },
        "UNITS": {}
    }

    try:
        xls = pd.ExcelFile(excel_file_path, engine='openpyxl')

        # Process standard sheets first
        if 'METADATA' in xls.sheet_names:
            metadata_df = pd.read_excel(xls, sheet_name='METADATA', header=None)
            output["METADATA"] = process_metadata_sheet(metadata_df)

        if 'VARIABLES' in xls.sheet_names:
            variables_df = pd.read_excel(xls, sheet_name='VARIABLES', header=None)
            formula_ws = formula_wb['VARIABLES'] if 'VARIABLES' in formula_wb.sheetnames else None
            value_ws = value_wb['VARIABLES'] if 'VARIABLES' in value_wb.sheetnames else None
            variables_result = process_variables_sheet(variables_df, formula_ws, value_ws)
            output["VARIABLES"] = variables_result

            # Formulas are now included directly in the variables, no need for separate section

        if 'CONNECTORS' in xls.sheet_names:
            connectors_df = pd.read_excel(xls, sheet_name='CONNECTORS', header=None)
            output["CONNECTORS"] = process_connectors_sheet(connectors_df, formula_wb)

        # Skip calculations sheet processing - not needed

        # Process UNIT sheets (any sheet not in the standard list and not ending with _(SS))
        standard_sheets = {'METADATA', 'VARIABLES', 'CONNECTORS', 'CALCULATIONS', 'README', 'CHANGELOG', 'SUPPL MATERIAL'}
        variables_ws = formula_wb['VARIABLES'] if 'VARIABLES' in formula_wb.sheetnames else None

        for sheet_name in xls.sheet_names:
            if sheet_name not in standard_sheets and not sheet_name.endswith('_(SS)'):
                unit_df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                formula_ws = formula_wb[sheet_name] if sheet_name in formula_wb.sheetnames else None
                value_ws = value_wb[sheet_name] if sheet_name in value_wb.sheetnames else None

                # NEW: detect the unit name from content (fallback to sheet-based pretty name)
                detected_unit_name = extract_unit_name(unit_df) or sheet_name.replace("_", " ").title()

                # Build the unit dict (this function already sets Unit Info -> Unit name accordingly)
                unit_dict = process_units_sheet(unit_df, sheet_name, formula_ws, value_ws, variables_ws)

                # Ensure the dict key is the detected unit name, and make it unique if needed
                unit_key = detected_unit_name
                if unit_key in output["UNITS"]:
                    base = unit_key
                    n = 2
                    while f"{base} ({n})" in output["UNITS"]:
                        n += 1
                    unit_key = f"{base} ({n})"

                output["UNITS"][unit_key] = unit_dict

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return

    # Write to JSON file
    try:
        with open(output_json_path, 'w', encoding='utf-8') as json_file:
            json.dump(output, json_file, indent=4, ensure_ascii=False)
        print(f"Successfully created JSON file at {output_json_path}")
    except Exception as e:
        print(f"Error writing JSON file: {e}")


def parse_excel_to_model_json(content: bytes) -> Dict[str, Any]:
    """Parse Excel content bytes and return the model JSON."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        output_path = f"{tmp_path}.json"
        excel_to_json(tmp_path, output_path)
        with open(output_path, "r", encoding="utf-8") as json_file:
            return json.load(json_file)
    finally:
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except Exception:
            pass
        try:
            Path(f"{tmp_path}.json").unlink(missing_ok=True)
        except Exception:
            pass


def parse_model_file(content: bytes, filename: str, *, is_admin: bool) -> Dict[str, Any]:
    """Parse an uploaded model file (Excel or JSON) into model_data."""
    suffix = Path(filename or "").suffix.lower()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded model file is empty",
        )

    if suffix == ".json":
        if not is_admin:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only admins can upload JSON model files",
            )
        try:
            return json.loads(content.decode("utf-8"))
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Failed to parse model file: {exc}",
            ) from exc

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        try:
            return parse_excel_to_model_json(content)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Failed to parse model file: {exc}",
            ) from exc

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unsupported file type. Use Excel (.xlsx/.xlsm/.xls). Admin can also upload .json.",
    )


def process_metadata_sheet(metadata_df: pd.DataFrame) -> Dict[str, str]:
    """Process the METADATA sheet into a dictionary."""
    metadata = {}
    for _, row in metadata_df.iterrows():
        if len(row) < 2 or pd.isna(row[0]):
            continue

        # Normalize key: replace newlines, collapse multiple spaces, uppercase
        key = ' '.join(str(row[0]).replace('\n', ' ').split()).upper()
        if len(row) < 2:
            value = ""
        elif pd.isna(row[1]):
            value = ""
        else:
            # Handle version as string
            if "VERSION" in key and pd.api.types.is_numeric_dtype(type(row[1])):
                value = str(row[1])
            else:
                value = str(row[1]).strip()
        metadata[key] = value

    # Map uppercase keys to output format
    return {
        "Model UID": metadata.get("MODEL UID", ""),
        "Model Name": metadata.get("MODEL NAME", ""),
        "Authors and Contributors": metadata.get("AUTHORS AND CONTRIBUTORS", ""),
        "Contact Info": metadata.get("CONTACT INFO", ""),
        "Creation Date": metadata.get("CREATION DATE", ""),
        "Last Updated": metadata.get("LAST UPDATED", ""),
        "Model Status": metadata.get("MODEL STATUS", ""),
        "Version": metadata.get("VERSION", ""),
        "Confidentiality": metadata.get("CONFIDENTIALITY", ""),
        "Sharing Layer": metadata.get("SHARING LAYER", ""),
        "Grade": metadata.get("GRADE", ""),
        "Reference Capacity": next((v for k, v in metadata.items() if "REFERENCE CAPACITY" in k), ""),
        "TRL": metadata.get("TRL", ""),
        "Keywords": metadata.get("KEYWORDS", ""),
        "Description": metadata.get("DESCRIPTION", ""),
        "DOI": metadata.get("DOI", ""),
        "Main Related Publication": metadata.get("MAIN RELATED PUBLICATION", ""),
        "References": metadata.get("REFERENCES", ""),
        "Software": metadata.get("SOFTWARE", ""),
        "Software File": metadata.get("SOFTWARE FILE", ""),
        "Block Flow Diagram": next((v for k, v in metadata.items() if "BLOCK FLOW DIAGRAM" in k), ""),
        "URL": metadata.get("URL", ""),
        "Supplementary Material": metadata.get("SUPPLEMENTARY MATERIAL", "")
    }


def process_variables_sheet(variables_df: pd.DataFrame, formula_ws=None, value_ws=None) -> List[Dict[str, Any]]:
    """Process the VARIABLES sheet with multiple sections, skipping EXAMPLE sections."""
    variables = []
    current_headers = []
    in_section = False
    skip_next_section = False  # Flag to skip EXAMPLE sections

    for row_idx, row in variables_df.iterrows():
        # Skip empty rows
        if all(pd.isna(cell) for cell in row):
            continue

        first_cell = str(row[0]).strip().upper() if len(row) > 0 and not pd.isna(row[0]) else ""

        # Detect EXAMPLE keyword - skip the next VARIABLES section
        if first_cell == 'EXAMPLE':
            skip_next_section = True
            continue

        # Check for VARIABLES keyword to start new section
        if first_cell == 'VARIABLES':
            if skip_next_section:
                # This is an example section, skip it
                skip_next_section = False
                in_section = False
                current_headers = []
            else:
                in_section = True
                current_headers = []  # Reset headers for new section
            continue

        # First non-empty row after VARIABLES is headers
        if in_section and not current_headers:
            current_headers = [str(cell).strip() for cell in row if not pd.isna(cell)]
            continue

        # Process data rows
        if in_section and current_headers:
            # Skip rows that don't start with a variable name
            if len(row) == 0 or pd.isna(row[0]):
                continue

            variable_data = {}
            for i, header in enumerate(current_headers):
                if i >= len(row):
                    break  # Skip if row doesn't have enough columns

                value = row[i]

                # Check for formula in any column BEFORE checking for nan
                formula, calc_value = get_formula_and_value(formula_ws, value_ws, variables_df, row_idx, i)

                # If we have a formula, use it even if pandas value is nan
                if formula:
                    # Get the calculated value - try calc_value first, then pandas value, then empty
                    final_value = calc_value
                    if final_value is None and not pd.isna(value):
                        final_value = value
                    if final_value is None:
                        final_value = ""

                    # Convert final_value to appropriate type
                    if final_value != "":
                        if pd.api.types.is_float(final_value):
                            final_value = float(final_value)
                        elif pd.api.types.is_integer(final_value):
                            final_value = int(final_value)
                        elif pd.api.types.is_bool(final_value):
                            final_value = bool(final_value)
                        else:
                            final_value = str(final_value).strip()

                    # Create readable formula with variable names
                    readable_formula = replace_cell_references_with_variable_names(formula, formula_ws, formula_ws)

                    variable_data[header] = {
                        "value": final_value,
                        "formula": readable_formula
                    }
                    continue

                # Always include all headers, set empty string if NaN
                if pd.isna(value):
                    variable_data[header] = ""
                    continue

                # Convert pandas/numpy types to native Python types
                if pd.api.types.is_float(value):
                    value = float(value)
                elif pd.api.types.is_integer(value):
                    value = int(value)
                elif pd.api.types.is_bool(value):
                    value = bool(value)
                else:
                    value = str(value).strip()

                variable_data[header] = value

            if variable_data:
                variables.append(variable_data)

    return variables


def is_valid_connector(connector_data: Dict[str, Any]) -> bool:
    """
    Validate that a connector entry has the minimum required fields.

    A valid connector must have:
    1. A non-empty NAME (ALIAS) field
    2. Either a VALUE field or at least 3 other meaningful fields
    """
    # Must have a name/alias
    name_field = connector_data.get("NAME (ALIAS)", "").strip()
    if not name_field or name_field in [",", "-", ""]:
        return False

    # Must have some substantial data beyond just name
    substantial_fields = 0
    for key, value in connector_data.items():
        if key == "NAME (ALIAS)":
            continue
        # Check if the field has meaningful content
        if value is not None and str(value).strip() not in ["", ",", "-"]:
            substantial_fields += 1

    # Need at least 2 other meaningful fields (e.g., TYPE, DIRECTION, VALUE, etc.)
    return substantial_fields >= 2


def process_connectors_sheet(connectors_df: pd.DataFrame, formula_wb=None) -> Dict[str, List[Dict[str, Any]]]:
    """Process the CONNECTORS sheet with EXTERNAL and INTERNAL sections."""
    connectors = {
        "EXTERNAL CONNECTORS": [],
        "INTERNAL CONNECTORS": []
    }
    current_section = None
    current_headers = []
    header_row_processed = False

    for idx, row in connectors_df.iterrows():
        # Skip empty rows
        if all(pd.isna(cell) for cell in row):
            continue

        # Check for section headers
        first_cell = str(row[0]).strip().upper() if len(row) > 0 else ""

        if "EXTERNAL CONNECTORS" in first_cell:
            current_section = "EXTERNAL CONNECTORS"
            header_row_processed = False
            continue
        elif "INTERNAL CONNECTORS" in first_cell:
            current_section = "INTERNAL CONNECTORS"
            header_row_processed = False
            continue

        # First non-empty row after section is headers (read only contiguous columns)
        if current_section and not header_row_processed:
            current_headers = []
            for cell in row:
                if pd.isna(cell):
                    break  # Stop at first empty column to avoid picking up side tables
                current_headers.append(str(cell).strip())
            header_row_processed = True
            continue

        # Process data rows (only if we have headers and we're in a section)
        if current_section and current_headers and header_row_processed:
            # Skip rows that don't start with a UID
            if len(row) == 0 or pd.isna(row[0]):
                continue

            # Skip rows that exactly match the headers (avoid including header row as data)
            if all(str(row[i]).strip() == current_headers[i] for i in range(min(len(row), len(current_headers)))):
                continue

            connector_data = {}
            for i, header in enumerate(current_headers):
                if i >= len(row):
                    break  # Skip if row doesn't have enough columns

                value = row[i]
                # Always include all headers, set empty string if NaN
                if pd.isna(value):
                    connector_data[header] = ""
                    continue

                # Check for formulas in connector cells
                formula = None
                calc_value = None
                if formula_wb and 'CONNECTORS' in formula_wb.sheetnames:
                    formula_ws = formula_wb['CONNECTORS']
                    variables_ws = formula_wb['VARIABLES'] if 'VARIABLES' in formula_wb.sheetnames else None
                    formula, calc_value = get_formula_from_unit_cell(formula_ws, None, idx, i, variables_ws)

                if formula:
                    # Convert formula to readable format
                    readable_formula = replace_cell_references_with_variable_names(formula, formula_ws, variables_ws)
                    final_value = calc_value if calc_value is not None else value

                    # Convert to appropriate type
                    if final_value != "" and not pd.isna(final_value):
                        if pd.api.types.is_float(final_value):
                            final_value = float(final_value)
                        elif pd.api.types.is_integer(final_value):
                            final_value = int(final_value)

                    connector_data[header] = {
                        "value": final_value,
                        "formula": readable_formula
                    }
                else:
                    # Convert values appropriately
                    if header in ["TEMPERATURE", "PRESSURE", "VAPOR FRACTION [MASS]"]:
                        try:
                            if str(value).upper() == "SAT":
                                value = "SAT"
                            else:
                                value = float(value)
                        except (ValueError, TypeError):
                            value = str(value).strip()
                    else:
                        value = str(value).strip()

                    connector_data[header] = value

            # Validate connector data before adding
            if connector_data and is_valid_connector(connector_data):
                connectors[current_section].append(connector_data)

    return connectors

def extract_unit_name(unit_df: pd.DataFrame) -> str:
    """
    Find a row where the FIRST COLUMN equals 'Unit name' (case-insensitive, parentheses allowed),
    and return the text in the FIRST COLUMN of the NEXT ROW (the cell directly below).
    """
    for i in range(len(unit_df) - 1):
        cell = unit_df.iloc[i, 0] if unit_df.shape[1] > 0 else None
        if pd.isna(cell):
            continue
        label = str(cell).strip().lower()
        # normalize: remove spaces and parentheses -> 'unitname'
        norm = re.sub(r'[\s()]+', '', label)
        if norm == 'unitname':
            below = unit_df.iloc[i + 1, 0]
            if not pd.isna(below) and str(below).strip():
                return str(below).strip()
    return ""

def process_units_sheet(unit_df: pd.DataFrame, sheet_name: str, formula_ws=None, value_ws=None, variables_ws=None) -> Dict[str, Any]:
    """
    Processes a UNIT sheet into a structured dictionary with all sections:
    - Unit Description
    - Unit Info (name, capacity multipliers)
    - Connectors
    - Heat Streams
    - Equipments
    """
    # NEW: prefer the cell below the 'Unit name' label; fallback to sheet name
    detected_unit_name = extract_unit_name(unit_df) or sheet_name.replace("_", " ").title()

    unit_data = {
        "Unit Description": "",
        "Unit Info": {
            "Unit name": detected_unit_name,
            "Minimum capacity multiplier": 1.0,
            "Maximum capacity multiplier": 1.0
        },
        "Connectors Description": "",
        "Connectors": [],
        "Heat Streams Description": "",
        "Heat Streams": [],
        "Equipments Description": "",
        "Equipments": {},
        "Operating Cost Description": "",
        "Operating Cost": {
            "Fixed Operating Cost [EUR/y]": 0.0,
            "Variable Operating Cost [EUR/h]": 0.0
        }
    }

    current_section = None
    headers = []
    skip_next = False
    unit_info_next_row = False
    unit_info_expected_idx = None

    for idx, row in unit_df.iterrows():
        if skip_next:
            skip_next = False
            continue

        # Skip empty rows
        if all(pd.isna(cell) for cell in row):
            continue

        # Get first cell value
        first_cell = str(row[0]).strip() if len(row) > 0 and not pd.isna(row[0]) else ""

        # ===== SECTION DETECTION =====
        # Unit Description
        if "unit description" in first_cell.lower():
            if idx + 1 < len(unit_df):
                next_row = unit_df.iloc[idx + 1]
                unit_data["Unit Description"] = str(next_row[0]).strip() if not pd.isna(next_row[0]) else ""
                skip_next = True
            continue

        # Unit Info - Header row (detect standard labels)
        elif (
            len(row) >= 3 and
            isinstance(row[0], str) and 'unit' in row[0].lower() and 'name' in row[0].lower() and
            isinstance(row[1], str) and 'capacity' in row[1].lower() and ('min' in row[1].lower() or 'minimum' in row[1].lower()) and
            isinstance(row[2], str) and 'capacity' in row[2].lower() and ('max' in row[2].lower() or 'maximum' in row[2].lower())
        ):
            unit_info_next_row = True
            unit_info_expected_idx = idx + 1
            continue

        # Unit Info - parse the row immediately following the header
        if unit_info_next_row and idx == unit_info_expected_idx:
            try:
                # Unit name in column 0
                if len(row) >= 1 and not pd.isna(row[0]) and str(row[0]).strip():
                    unit_data["Unit Info"]["Unit name"] = str(row[0]).strip()

                # Min capacity in column 1
                if len(row) >= 2 and not pd.isna(row[1]):
                    # Try to read formula/value from the workbook; fallback to the raw cell value
                    f, cv = get_formula_from_unit_cell(formula_ws, value_ws, idx, 1, variables_ws) if formula_ws else (None, None)
                    if f:
                        unit_data["Unit Info"]["Minimum capacity multiplier"] = {
                            "value": cv if cv is not None else safe_float(row[1]),
                            "formula": replace_cell_references_with_variable_names(f, formula_ws, variables_ws)
                        }
                    else:
                        unit_data["Unit Info"]["Minimum capacity multiplier"] = safe_float(row[1])

                # Max capacity in column 2
                if len(row) >= 3 and not pd.isna(row[2]):
                    f, cv = get_formula_from_unit_cell(formula_ws, value_ws, idx, 2, variables_ws) if formula_ws else (None, None)
                    if f:
                        unit_data["Unit Info"]["Maximum capacity multiplier"] = {
                            "value": cv if cv is not None else safe_float(row[2]),
                            "formula": replace_cell_references_with_variable_names(f, formula_ws, variables_ws)
                        }
                    else:
                        unit_data["Unit Info"]["Maximum capacity multiplier"] = safe_float(row[2])
            except Exception:
                pass
            finally:
                unit_info_next_row = False
                unit_info_expected_idx = None
            continue
        
        # Connectors Description
        elif "connectors description" in first_cell.lower():
            if idx + 1 < len(unit_df):
                next_row = unit_df.iloc[idx + 1]
                unit_data["Connectors Description"] = str(next_row[0]).strip() if not pd.isna(next_row[0]) else ""
                skip_next = True
            continue

        # Connectors Data
        elif first_cell.upper() == "CONNECTORS":
            current_section = "Connectors"
            if idx + 1 < len(unit_df):
                headers = [str(c).strip() for c in unit_df.iloc[idx + 1] if not pd.isna(c)]
                skip_next = True
            continue

        # Heat Streams Description
        elif "heat streams description" in first_cell.lower():
            if idx + 1 < len(unit_df):
                next_row = unit_df.iloc[idx + 1]
                unit_data["Heat Streams Description"] = str(next_row[0]).strip() if not pd.isna(next_row[0]) else ""
                skip_next = True
            continue

        # Heat Streams Data
        elif first_cell.upper() == "HEAT STREAMS":
            current_section = "Heat Streams"
            if idx + 1 < len(unit_df):
                headers = [str(c).strip() for c in unit_df.iloc[idx + 1] if not pd.isna(c)]
                skip_next = True
            continue

        # Equipments Description
        elif "equipments description" in first_cell.lower():
            if idx + 1 < len(unit_df):
                next_row = unit_df.iloc[idx + 1]
                unit_data["Equipments Description"] = str(next_row[0]).strip() if not pd.isna(next_row[0]) else ""
                skip_next = True
            continue

        # Equipments Data
        elif first_cell.upper() == "EQUIPMENTS":
            current_section = "Equipments"
            if idx + 1 < len(unit_df):
                headers = [str(c).strip() for c in unit_df.iloc[idx + 1] if not pd.isna(c)]
                skip_next = True
            continue

        # Operating Cost Description
        elif "operating cost description" in first_cell.lower():
            if idx + 1 < len(unit_df):
                next_row = unit_df.iloc[idx + 1]
                unit_data["Operating Cost Description"] = str(next_row[0]).strip() if not pd.isna(next_row[0]) else ""
                skip_next = True
            continue

        # Operating Cost Data - detect the headers row
        elif ("fixed operating cost" in first_cell.lower() and
              len(row) > 1 and "variable operating cost" in str(row[1]).lower()):
            current_section = "Operating Cost"
            headers = [str(cell).strip() for cell in row[:2] if not pd.isna(cell)]
            continue

        # ===== DATA PROCESSING =====
        if current_section and headers:
            # Skip rows that exactly match headers
            if all(i < len(row) and str(row[i]).strip() == h for i, h in enumerate(headers)):
                continue

            # Process Connectors
            if current_section == "Connectors":
                connector = process_connector_row(row, headers, formula_ws, value_ws, variables_ws, idx)
                if connector:
                    unit_data["Connectors"].append(connector)

            # Process Heat Streams
            elif current_section == "Heat Streams":
                heat_stream = process_heat_stream_row(row, headers, formula_ws, value_ws, variables_ws, idx)
                if heat_stream:
                    unit_data["Heat Streams"].append(heat_stream)

            # Process Equipments
            elif current_section == "Equipments":
                process_equipment_row(row, headers, unit_data["Equipments"], formula_ws, value_ws, variables_ws, idx)

            # Process Operating Cost
            elif current_section == "Operating Cost":
                if len(row) >= 2:
                    unit_data["Operating Cost"]["Fixed Operating Cost [EUR/y]"] = safe_float(row[0], 0.0)
                    unit_data["Operating Cost"]["Variable Operating Cost [EUR/h]"] = safe_float(row[1], 0.0)

    return unit_data


def safe_float(value, default=1.0):
    """Convert to float safely with default fallback."""
    try:
        return float(value) if not pd.isna(value) else default
    except (ValueError, TypeError):
        return default


def get_formula_from_unit_cell(formula_ws, value_ws, row_idx, col_idx, variables_ws=None, match_value=None, match_col=1):
    """Get formula and calculated value from a specific unit sheet cell."""
    formula = None
    calculated_value = None

    if not formula_ws:
        return formula, calculated_value

    try:
        target_row = row_idx + 1
        if match_value:
            # Try to locate the row by matching a value in the specified column (default col 1 / A)
            for excel_row in range(1, formula_ws.max_row + 1):
                cell_val = formula_ws.cell(row=excel_row, column=match_col).value
                if str(cell_val).strip() == str(match_value).strip():
                    target_row = excel_row
                    break

        # Get calculated value first (even if no formula)
        if value_ws:
            value_cell = value_ws.cell(row=target_row, column=col_idx + 1)
            calculated_value = value_cell.value

        # Get the formula from the specific cell
        formula_cell = formula_ws.cell(row=target_row, column=col_idx + 1)  # Convert to 1-based
        if formula_cell.value and str(formula_cell.value).startswith('='):
            formula = str(formula_cell.value)
            if calculated_value is None:
                evaluator = FormulaEvaluator(formula_ws)
                cell_ref = f"{get_column_letter(col_idx + 1)}{target_row}"
                calculated_value = evaluator.get_cell_value(cell_ref)

        return formula, calculated_value
    except:
        return formula, calculated_value


def process_connector_row(row, headers, formula_ws=None, value_ws=None, variables_ws=None, row_idx=None):
    """Process a single row of connectors data."""
    connector = {}
    for i, header in enumerate(headers):
        if i >= len(row):
            connector[header] = ""
            continue

        value = row[i]
        # Always include all headers, set empty string if NaN
        if pd.isna(value):
            connector[header] = ""
            continue

        # Check for formulas in this cell
        formula = None
        calc_value = None
        if formula_ws and row_idx is not None:
            formula, calc_value = get_formula_from_unit_cell(formula_ws, value_ws, row_idx, i, variables_ws)

        if formula:
            # Convert formula to readable format
            readable_formula = replace_cell_references_with_variable_names(formula, formula_ws, variables_ws)
            final_value = calc_value if calc_value is not None else value

            # Convert final_value to appropriate type
            if final_value != "" and not pd.isna(final_value):
                if pd.api.types.is_float(final_value):
                    final_value = float(final_value)
                elif pd.api.types.is_integer(final_value):
                    final_value = int(final_value)

            connector[header] = {
                "value": final_value,
                "formula": readable_formula
            }
        else:
            # Handle unit fields - they should always be strings (like "C", "bar", "kg/h")
            if "unit" in header.lower():
                connector[header] = str(value).strip()
            # Convert numeric fields to floats
            elif any(x in header.lower() for x in ["value", "temperature", "pressure", "fraction"]):
                text_value = str(value).strip()
                if text_value in {"", "-", "–", "—"}:
                    connector[header] = text_value  # preserve dash placeholders
                    continue
                connector[header] = safe_float(value)
            else:
                connector[header] = value

    # Validate connector before returning
    if connector and is_valid_connector(connector):
        return connector
    else:
        return None


def process_heat_stream_row(row, headers, formula_ws=None, value_ws=None, variables_ws=None, row_idx=None):
    """Process a single row of heat streams data."""
    stream = {}
    stream_name = None
    if len(row) > 0 and not pd.isna(row[0]):
        stream_name = str(row[0]).strip()
    for i, header in enumerate(headers):
        if i >= len(row):
            stream[header] = ""
            continue

        value = row[i]
        # Check for formulas in this cell (before NaN short-circuit)
        formula = None
        calc_value = None
        if formula_ws and row_idx is not None:
            formula, calc_value = get_formula_from_unit_cell(
                formula_ws,
                value_ws,
                row_idx,
                i,
                variables_ws,
                match_value=stream_name,
                match_col=1
            )

        if formula:
            # Convert formula to readable format
            readable_formula = replace_cell_references_with_variable_names(formula, formula_ws, variables_ws)
            final_value = calc_value if calc_value is not None else value

            # Convert final_value to appropriate type
            if final_value != "" and not pd.isna(final_value):
                if pd.api.types.is_float(final_value):
                    final_value = float(final_value)
                elif pd.api.types.is_integer(final_value):
                    final_value = int(final_value)

            stream[header] = {
                "value": final_value,
                "formula": readable_formula
            }
        else:
            # Always include all headers, set empty string if NaN (unless we have a calculated value)
            if pd.isna(value):
                if calc_value is not None and not pd.isna(calc_value):
                    value = calc_value
                else:
                    stream[header] = ""
                    continue
            # Handle unit fields - they should always be strings (like "C", "kW")
            if "unit" in header.lower():
                stream[header] = str(value).strip()
            # Convert numeric VALUE fields to floats (matches TEMPERATURE VALUE, ENTHALPY FLOW VALUE, etc.)
            elif "value" in header.lower():
                text_value = str(value).strip()
                if text_value in {"", "-", "–", "—"}:
                    stream[header] = text_value
                    continue
                stream[header] = safe_float(value)
            else:
                stream[header] = value

    return stream if stream else None


def process_equipment_row(row, headers, equipments, formula_ws=None, value_ws=None, variables_ws=None, row_idx=None):
    """Process equipment rows with hierarchical structure.

    New v6 format columns: NAME(0), TYPE(1), SUBTYPE(2), REFERENCE(3), PARAMS(4), VALUE(5), UNIT(6)
    """
    if len(row) < 1:
        return

    # New equipment entry - when first column has equipment name
    if len(row) >= 2 and not pd.isna(row[0]) and not pd.isna(row[1]):
        eq_name = str(row[0]).strip()
        eq_type = str(row[1]).strip()
        eq_subtype = str(row[2]).strip() if len(row) > 2 and not pd.isna(row[2]) else ""
        eq_reference = str(row[3]).strip() if len(row) > 3 and not pd.isna(row[3]) else ""
        eq_key = eq_name

        if eq_key not in equipments:
            equipments[eq_key] = {
                "Type": eq_type,
                "Subtype": eq_subtype,
                "Reference": eq_reference,
                "Parameters": []
            }

        # Add parameters if available (PARAMS at index 4, VALUE at index 5, UNIT at index 6)
        if len(row) >= 7 and not pd.isna(row[4]):
            # Check for formula in the Value column (index 5)
            value = row[5] if len(row) > 5 else None
            formula = None
            calc_value = None

            if formula_ws and row_idx is not None and len(row) > 5:
                formula, calc_value = get_formula_from_unit_cell(formula_ws, value_ws, row_idx, 5, variables_ws)

            if formula:
                # Convert formula to readable format
                readable_formula = replace_cell_references_with_variable_names(formula, formula_ws, variables_ws)
                final_value = calc_value if calc_value is not None else value

                # Convert to appropriate type
                if final_value != "" and not pd.isna(final_value):
                    if pd.api.types.is_float(final_value):
                        final_value = float(final_value)
                    elif pd.api.types.is_integer(final_value):
                        final_value = int(final_value)

                param_value = {
                    "value": final_value,
                    "formula": readable_formula
                }
            else:
                param_value = safe_convert(value) if value is not None else ""

            param = {
                "Parameter": str(row[4]).strip() if not pd.isna(row[4]) else "",
                "Value": param_value,
                "Unit": str(row[6]).strip() if len(row) > 6 and not pd.isna(row[6]) else ""
            }
            equipments[eq_key]["Parameters"].append(param)

    # Continuation row for parameters - first 4 columns empty, data in PARAMS(4), VALUE(5), UNIT(6)
    elif equipments and len(row) >= 5 and all(pd.isna(row[i]) or str(row[i]).strip() == "" for i in range(4)) and not pd.isna(row[4]):
        last_eq = list(equipments.keys())[-1]

        # Check for formula in the Value column (index 5)
        value = row[5] if len(row) > 5 else None
        formula = None
        calc_value = None

        if formula_ws and row_idx is not None and len(row) > 5:
            formula, calc_value = get_formula_from_unit_cell(formula_ws, value_ws, row_idx, 5, variables_ws)

        if formula:
            # Convert formula to readable format
            readable_formula = replace_cell_references_with_variable_names(formula, formula_ws, variables_ws)
            final_value = calc_value if calc_value is not None else value

            # Convert to appropriate type
            if final_value != "" and not pd.isna(final_value):
                if pd.api.types.is_float(final_value):
                    final_value = float(final_value)
                elif pd.api.types.is_integer(final_value):
                    final_value = int(final_value)

            param_value = {
                "value": final_value,
                "formula": readable_formula
            }
        else:
            param_value = safe_convert(value) if value is not None else ""
            
        param = {
            "Parameter": str(row[4]).strip() if not pd.isna(row[4]) else "",
            "Value": param_value,
            "Unit": str(row[6]).strip() if len(row) > 6 and not pd.isna(row[6]) else ""
        }
        equipments[last_eq]["Parameters"].append(param)


def safe_convert(value):
    """Smart value conversion for equipment parameters."""
    if pd.isna(value):
        return ""
    try:
        return float(value) if str(value).replace('.', '', 1).isdigit() else str(value).strip()
    except:
        return str(value).strip()


def extract_value_with_formula(formula_ws, value_ws, variables_ws, row_idx, col_idx):
    """Extract value from Excel cell, handling formulas if present"""
    # Check for formula first
    formula, calc_value = get_formula_from_unit_cell(formula_ws, value_ws, row_idx, col_idx, variables_ws)

    if formula:
        # Convert formula to readable format
        readable_formula = replace_cell_references_with_variable_names(formula, formula_ws, variables_ws)
        final_value = calc_value if calc_value is not None else 0

        return {
            "value": final_value,
            "formula": readable_formula
        }
    else:
        # No formula, just return the raw value
        return safe_float(calc_value if calc_value is not None else 0)



if __name__ == "__main__":
    import argparse

    script_dir = Path(__file__).resolve().parent

    arg_parser = argparse.ArgumentParser(description="Parse Excel v6 files to JSON")
    arg_parser.add_argument(
        "--all", action="store_true",
        help=f"Process all .xlsx files in {V6_EXCEL_DIR}/ and write JSON to {V6_JSON_DIR}/"
    )
    arg_parser.add_argument("--input", type=str, default=None, help="Single Excel input file")
    arg_parser.add_argument("--output", type=str, default=None, help="Single JSON output file")
    args = arg_parser.parse_args()

    if args.all:
        excel_dir = script_dir / V6_EXCEL_DIR
        json_dir = script_dir / V6_JSON_DIR
        json_dir.mkdir(parents=True, exist_ok=True)

        xlsx_files = list(excel_dir.glob("*.xlsx")) + list(excel_dir.glob("*.xlsm"))
        if not xlsx_files:
            print(f"No Excel files found in {excel_dir}")
        else:
            for xlsx_path in sorted(xlsx_files):
                out_path = json_dir / (xlsx_path.stem + ".json")
                print(f"Processing {xlsx_path.name} -> {out_path.name}")
                excel_to_json(str(xlsx_path), str(out_path))
    else:
        input_excel = Path(args.input) if args.input else Path(INPUT_EXCEL_NAME)
        output_json = Path(args.output) if args.output else Path(OUTPUT_JSON_NAME)

        resolved_input = input_excel if input_excel.is_absolute() else (script_dir / input_excel)
        resolved_output = output_json if output_json.is_absolute() else (script_dir / output_json)

        resolved_output.parent.mkdir(parents=True, exist_ok=True)
        excel_to_json(str(resolved_input), str(resolved_output))
